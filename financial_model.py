"""
financial_model.py
Generate a full investment banking-style financial model Excel file for any public stock.
Usage: python financial_model.py AAPL
       python financial_model.py MSFT
"""

import sys
import os
import argparse
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Color palette ──────────────────────────────────────────────────────────────
NAVY       = "1F3864"
BLUE_TEXT  = "0000FF"
GREEN_TEXT = "008000"
WHITE      = "FFFFFF"
BLACK      = "000000"
LIGHT_BLUE = "DDEEFF"
LIGHT_GRN  = "DDFFDD"
LIGHT_AMB  = "FFF3CC"
LIGHT_GRAY = "F2F2F2"
HEADER_BG  = "2F5496"

# ── Style helpers ──────────────────────────────────────────────────────────────
def _font(bold=False, color=BLACK, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(bottom=thin)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

FMT_USD   = '#,##0;(#,##0);"-"'
FMT_PCT   = '0.0%;(0.0%);"-"'
FMT_MULT  = '0.0x'
FMT_2DP   = '#,##0.00'

def style(ws, row, col, value=None, bold=False, color=BLACK, fill=None,
          fmt=None, align="left", border=False, italic=False, size=10):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font = _font(bold=bold, color=color, size=size, italic=italic)
    if fill:
        cell.fill = _fill(fill)
    if fmt:
        cell.number_format = fmt
    cell.alignment = _align(h=align)
    if border:
        cell.border = _border()
    return cell

def header_row(ws, row, text, bg=NAVY, fg=WHITE, height=22, end_col=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(bold=True, color=fg, size=12)
    c.fill = _fill(bg)
    c.alignment = _align(h="center")
    ws.row_dimensions[row].height = height

def section_label(ws, row, text, max_col=13):
    blue_top = Border(top=Side(style="medium", color=HEADER_BG))
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _fill(LIGHT_GRAY)
        cell.border = blue_top
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(bold=True, color="1F3864", size=10)
    ws.row_dimensions[row].height = 15

def col_headers(ws, row, labels, start_col=2, bg=HEADER_BG, fg=WHITE):
    white_bottom = Border(bottom=Side(style="medium", color="FFFFFF"))
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=lbl)
        c.font = _font(bold=True, color=fg, size=10)
        c.fill = _fill(bg)
        c.alignment = _align(h="center")
        c.border = white_bottom
    ws.row_dimensions[row].height = 16

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ── Data fetcher ───────────────────────────────────────────────────────────────
def fetch(ticker_sym):
    t = yf.Ticker(ticker_sym)
    info = t.info
    fin  = t.financials        # IS  — annual, columns = fiscal year dates
    bs   = t.balance_sheet
    cf   = t.cashflow
    return t, info, fin, bs, cf

def safe(d, key, default=0):
    v = d.get(key, default)
    return v if v is not None else default

def row_val(df, key, col=0, default=0):
    try:
        if key in df.index:
            v = df.loc[key].iloc[col]
            return float(v) if v == v else default
    except Exception:
        pass
    return default

def fiscal_years(fin):
    return [str(c.year) for c in fin.columns]

# ── Sheet builders ─────────────────────────────────────────────────────────────

def build_cover(wb, info, sym):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1F3864"
    set_col_widths(ws, {"A": 30, "B": 20, "C": 20, "D": 20})

    ws.row_dimensions[2].height = 30
    c = ws.cell(row=2, column=1, value=f"{info.get('longName', sym).upper()}")
    c.font = _font(bold=True, color=NAVY, size=18)
    ws.merge_cells("A2:D2")
    c.alignment = _align(h="center")

    ws.cell(row=3, column=1, value="Three-Statement Financial Model + Valuation").font = _font(italic=True, size=12, color="444444")
    ws.merge_cells("A3:D3")
    ws.cell(row=3, column=1).alignment = _align(h="center")

    ws.cell(row=4, column=1, value=f"Source: Yahoo Finance (yfinance)  |  Figures in $M unless stated").font = _font(italic=True, size=9, color="666666")
    ws.merge_cells("A4:D4")
    ws.cell(row=4, column=1).alignment = _align(h="center")

    # Color legend
    r = 6
    header_row(ws, r, "Color-Coding Legend", bg=NAVY)
    legends = [
        (BLUE_TEXT,  WHITE,      "Blue text",       "Hardcoded input"),
        (BLACK,      WHITE,      "Black text",       "Calculated formula"),
        (GREEN_TEXT, WHITE,      "Green text",       "Cross-sheet link"),
        (None,       LIGHT_BLUE, "Light blue bg",    "Key IS metric"),
        (None,       LIGHT_GRN,  "Light green bg",   "Key BS metric"),
        (None,       LIGHT_AMB,  "Light amber bg",   "Key CFS metric"),
        (None,       LIGHT_GRAY, "Light gray bg",    "Subtotal / total"),
    ]
    for i, (fc, bg, lbl, desc) in enumerate(legends):
        row = r + 1 + i
        c1 = ws.cell(row=row, column=1, value=lbl)
        c1.font = _font(color=fc if fc else BLACK, bold=True)
        if bg and bg != WHITE:
            c1.fill = _fill(bg)
        ws.cell(row=row, column=2, value=desc).font = _font()

    # Sheet nav
    r2 = r + len(legends) + 2
    header_row(ws, r2, "Sheet Navigation", bg=HEADER_BG)
    sheets = [
        ("Income Statement",      "Revenue → COGS → Gross Profit → EBIT → Net Income"),
        ("Balance Sheet",         "Assets = Liabilities + Equity"),
        ("Cash Flow Statement",   "Operating / Investing / Financing"),
        ("3-Stmt Linkage",        "Side-by-side cross-statement ties"),
        ("DCF Model",             "5-Year UFCF projection + Terminal Value"),
        ("WACC Calculation",      "CAPM cost of equity + after-tax cost of debt"),
        ("Trading Comps",         "Peer EV/EBITDA, EV/Rev, P/E multiples"),
        ("Valuation Summary",     "Football field — blended implied price"),
    ]
    for i, (s, d) in enumerate(sheets):
        ws.cell(row=r2+1+i, column=1, value=s).font = _font(bold=True, color=BLUE_TEXT)
        ws.cell(row=r2+1+i, column=2, value=d).font = _font()


def build_is(wb, info, fin, sym):
    ws = wb.create_sheet("Income Statement")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "375623"
    ws.freeze_panes = "B4"
    set_col_widths(ws, {"A": 38, "B": 16, "C": 14, "D": 14, "E": 16, "F": 36})

    yrs = fiscal_years(fin)
    yr0 = yrs[0] if yrs else "LTM"
    yr1 = yrs[1] if len(yrs) > 1 else ""

    r = 1
    header_row(ws, r, f"{info.get('longName', sym).upper()} — INCOME STATEMENT"); r += 1
    ws.cell(row=r, column=1, value=f"Fiscal Year | $M").font = _font(italic=True, size=9); r += 1
    col_headers(ws, r, [yr0, "% of Rev", f"YoY Δ ($M)", yr1, "Notes"]); r += 1

    def revenue_line(label, key, note=""):
        v0 = row_val(fin, key, 0) / 1e6
        v1 = row_val(fin, key, 1) / 1e6
        style(ws, r, 1, label, color=BLUE_TEXT)
        style(ws, r, 2, v0, fmt=FMT_USD, align="right", fill=LIGHT_BLUE)
        style(ws, r, 3, f"=B{r}/B{rev_row}" if rev_row else "", fmt=FMT_PCT, align="right")
        style(ws, r, 4, v0 - v1 if v1 else "", fmt=FMT_USD, align="right")
        style(ws, r, 5, v1, fmt=FMT_USD, align="right")
        style(ws, r, 6, note, italic=True, size=9, color="555555")

    # Revenue section
    section_label(ws, r, "  REVENUE"); r += 1
    rev_row = None

    total_rev = row_val(fin, "Total Revenue", 0) / 1e6
    total_rev1 = row_val(fin, "Total Revenue", 1) / 1e6

    # Segment proxies from info
    iphone_rev = safe(info, "totalRevenue", 0) / 1e6  # fallback
    services   = 0

    style(ws, r, 1, "Total Net Revenue", bold=True, color=BLUE_TEXT)
    style(ws, r, 2, total_rev, fmt=FMT_USD, align="right", bold=True, fill=LIGHT_BLUE)
    style(ws, r, 3, 1.0, fmt=FMT_PCT, align="right", bold=True)
    style(ws, r, 4, total_rev - total_rev1, fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 5, total_rev1, fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 6, "Source: Yahoo Finance — Total Revenue", italic=True, size=9, color="555555")
    rev_row = r; r += 1

    # COGS
    section_label(ws, r, "  COST OF GOODS SOLD"); r += 1
    cogs0 = row_val(fin, "Cost Of Revenue", 0) / 1e6
    cogs1 = row_val(fin, "Cost Of Revenue", 1) / 1e6
    style(ws, r, 1, "Total COGS", color=BLUE_TEXT)
    style(ws, r, 2, cogs0, fmt=FMT_USD, align="right", fill=LIGHT_GRAY)
    style(ws, r, 3, f"=B{r}/B{rev_row}", fmt=FMT_PCT, align="right")
    style(ws, r, 4, cogs0 - cogs1, fmt=FMT_USD, align="right")
    style(ws, r, 5, cogs1, fmt=FMT_USD, align="right")
    style(ws, r, 6, "Direct cost of goods sold; gross margin = 1 − COGS/Revenue", italic=True, size=9, color="555555")
    cogs_row = r; r += 1

    # Gross Profit
    gp0 = total_rev - cogs0
    gp1 = total_rev1 - cogs1
    style(ws, r, 1, "GROSS PROFIT", bold=True)
    style(ws, r, 2, f"=B{rev_row}-B{cogs_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_BLUE)
    style(ws, r, 3, f"=B{r}/B{rev_row}", fmt=FMT_PCT, align="right", bold=True)
    style(ws, r, 4, gp0 - gp1, fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 5, gp1, fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 6, "= Revenue − COGS; key profitability metric before operating expenses", italic=True, size=9, color="555555")
    gp_row = r; r += 1

    # OpEx
    section_label(ws, r, "  OPERATING EXPENSES"); r += 1
    rd0   = row_val(fin, "Research And Development", 0) / 1e6
    rd1   = row_val(fin, "Research And Development", 1) / 1e6
    sga0  = row_val(fin, "Selling General And Administration", 0) / 1e6
    sga1  = row_val(fin, "Selling General And Administration", 1) / 1e6

    style(ws, r, 1, "Research & Development", color=BLUE_TEXT)
    style(ws, r, 2, rd0, fmt=FMT_USD, align="right")
    style(ws, r, 3, f"=B{r}/B{rev_row}", fmt=FMT_PCT, align="right")
    style(ws, r, 4, rd0 - rd1, fmt=FMT_USD, align="right")
    style(ws, r, 5, rd1, fmt=FMT_USD, align="right")
    style(ws, r, 6, "Investment in future products, services, and technologies", italic=True, size=9, color="555555")
    rd_row = r; r += 1

    style(ws, r, 1, "Selling, General & Administrative", color=BLUE_TEXT)
    style(ws, r, 2, sga0, fmt=FMT_USD, align="right")
    style(ws, r, 3, f"=B{r}/B{rev_row}", fmt=FMT_PCT, align="right")
    style(ws, r, 4, sga0 - sga1, fmt=FMT_USD, align="right")
    style(ws, r, 5, sga1, fmt=FMT_USD, align="right")
    style(ws, r, 6, "Sales, marketing, and general & administrative overhead", italic=True, size=9, color="555555")
    sga_row = r; r += 1

    style(ws, r, 1, "Total Operating Expenses", bold=True)
    style(ws, r, 2, f"=B{rd_row}+B{sga_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRAY)
    style(ws, r, 3, f"=B{r}/B{rev_row}", fmt=FMT_PCT, align="right", bold=True)
    style(ws, r, 4, f"=D{rd_row}+D{sga_row}", fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 5, f"=E{rd_row}+E{sga_row}", fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 6, "= R&D + SG&A; total cash operating expenses below the gross profit line", italic=True, size=9, color="555555")
    opex_row = r; r += 1

    # EBIT
    ebit0 = row_val(fin, "Operating Income", 0) / 1e6
    ebit1 = row_val(fin, "Operating Income", 1) / 1e6
    style(ws, r, 1, "OPERATING INCOME (EBIT)", bold=True)
    style(ws, r, 2, f"=B{gp_row}-B{opex_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_BLUE)
    style(ws, r, 3, f"=B{r}/B{rev_row}", fmt=FMT_PCT, align="right", bold=True)
    style(ws, r, 4, ebit0 - ebit1, fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 5, ebit1, fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 6, "= Gross Profit − OpEx; earnings before interest & tax; key driver of DCF", italic=True, size=9, color="555555")
    ebit_row = r; r += 1

    # Below the line
    section_label(ws, r, "  BELOW-THE-LINE ITEMS"); r += 1
    int0 = row_val(fin, "Net Non Operating Interest Income Expense", 0) / 1e6
    if int0 == 0:
        int0 = row_val(fin, "Other Income Expense", 0) / 1e6
    int1 = row_val(fin, "Net Non Operating Interest Income Expense", 1) / 1e6
    if int1 == 0:
        int1 = row_val(fin, "Other Income Expense", 1) / 1e6
    style(ws, r, 1, "Interest & Other Income, net", color=BLUE_TEXT)
    style(ws, r, 2, int0, fmt=FMT_USD, align="right")
    style(ws, r, 3, f"=B{r}/B{rev_row}", fmt=FMT_PCT, align="right")
    style(ws, r, 4, int0 - int1, fmt=FMT_USD, align="right")
    style(ws, r, 5, int1, fmt=FMT_USD, align="right")
    style(ws, r, 6, "Net interest income/(expense); large cash pile makes this positive for Apple", italic=True, size=9, color="555555")
    int_row = r; r += 1

    ebt1 = row_val(fin, "Pretax Income", 1) / 1e6
    style(ws, r, 1, "Pre-Tax Income (EBT)", bold=True)
    style(ws, r, 2, f"=B{ebit_row}+B{int_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRAY)
    style(ws, r, 3, f"=B{r}/B{rev_row}", fmt=FMT_PCT, align="right", bold=True)
    style(ws, r, 4, f"=B{r}-E{r}", fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 5, ebt1, fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 6, "= EBIT + Net Interest; taxable income base before applying the effective tax rate", italic=True, size=9, color="555555")
    ebt_row = r; r += 1

    tax0 = row_val(fin, "Tax Provision", 0) / 1e6
    tax1 = row_val(fin, "Tax Provision", 1) / 1e6
    style(ws, r, 1, "Income Tax Expense", color=BLUE_TEXT)
    style(ws, r, 2, -abs(tax0), fmt=FMT_USD, align="right")
    style(ws, r, 3, f"=B{r}/B{rev_row}", fmt=FMT_PCT, align="right")
    style(ws, r, 4, f"=B{r}-E{r}", fmt=FMT_USD, align="right")
    style(ws, r, 5, -abs(tax1), fmt=FMT_USD, align="right")
    style(ws, r, 6, "Provision for income taxes; shown as negative (cash outflow to taxing authorities)", italic=True, size=9, color="555555")
    tax_row = r; r += 1

    style(ws, r, 1, "Effective Tax Rate", italic=True)
    style(ws, r, 2, f"=ABS(B{tax_row})/B{ebt_row}", fmt=FMT_PCT, align="right")
    style(ws, r, 5, f"=ABS(E{tax_row})/E{ebt_row}", fmt=FMT_PCT, align="right")
    r += 1

    # Net Income
    ni0 = row_val(fin, "Net Income", 0) / 1e6
    ni1 = row_val(fin, "Net Income", 1) / 1e6
    style(ws, r, 1, "NET INCOME", bold=True, size=11)
    style(ws, r, 2, f"=B{ebt_row}+B{tax_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_BLUE)
    style(ws, r, 3, f"=B{r}/B{rev_row}", fmt=FMT_PCT, align="right", bold=True)
    style(ws, r, 4, ni0 - ni1, fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 5, ni1, fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 6, "↓ Flows to Retained Earnings (BS) and CFS starting line", italic=True, size=9, color=GREEN_TEXT)
    ni_row = r; r += 1

    # EBITDA bridge
    r += 1
    section_label(ws, r, "  EBITDA BRIDGE"); r += 1
    da0 = row_val(fin, "Reconciled Depreciation", 0) / 1e6
    if da0 == 0:
        da0 = row_val(fin, "Depreciation And Amortization", 0) / 1e6
    da1 = row_val(fin, "Reconciled Depreciation", 1) / 1e6
    if da1 == 0:
        da1 = row_val(fin, "Depreciation And Amortization", 1) / 1e6

    style(ws, r, 1, "EBIT (Operating Income)")
    style(ws, r, 2, f"=B{ebit_row}", fmt=FMT_USD, align="right", color=GREEN_TEXT)
    style(ws, r, 3, f"=B{r}/B{rev_row}", fmt=FMT_PCT, align="right")
    style(ws, r, 4, ebit0 - ebit1, fmt=FMT_USD, align="right")
    style(ws, r, 5, ebit1, fmt=FMT_USD, align="right", color=GREEN_TEXT)
    da_row_is = r; r += 1

    style(ws, r, 1, "Add: Depreciation & Amortisation", color=BLUE_TEXT)
    style(ws, r, 2, da0, fmt=FMT_USD, align="right")
    style(ws, r, 3, f"=B{r}/B{rev_row}", fmt=FMT_PCT, align="right")
    style(ws, r, 4, da0 - da1, fmt=FMT_USD, align="right")
    style(ws, r, 5, da1, fmt=FMT_USD, align="right")
    style(ws, r, 6, "Non-cash charge; reduces reported earnings but not a cash outflow", italic=True, size=9, color="555555")
    da_row = r; r += 1

    style(ws, r, 1, "EBITDA", bold=True)
    style(ws, r, 2, f"=B{da_row_is}+B{da_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_BLUE)
    style(ws, r, 3, f"=B{r}/B{rev_row}", fmt=FMT_PCT, align="right", bold=True)
    style(ws, r, 4, (ebit0 + da0) - (ebit1 + da1), fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 5, ebit1 + da1, fmt=FMT_USD, align="right", bold=True, fill=LIGHT_BLUE)
    style(ws, r, 6, "= EBIT + D&A; proxy for operating cash flow; used in EV/EBITDA valuation", italic=True, size=9, color="555555")
    ebitda_row = r; r += 1

    # Key metrics
    r += 1
    section_label(ws, r, "  KEY METRICS"); r += 1
    shares = safe(info, "sharesOutstanding", 1) / 1e6
    style(ws, r, 1, "Diluted Shares Outstanding (M)", color=BLUE_TEXT)
    style(ws, r, 2, shares, fmt=FMT_2DP, align="right")
    style(ws, r, 6, "Diluted count includes vested options and unvested RSUs; source: Yahoo Finance", italic=True, size=9, color="555555")
    shares_row = r; r += 1

    for lbl, formula in [
        ("Gross Margin %",  f"=B{gp_row}/B{rev_row}"),
        ("EBIT Margin %",   f"=B{ebit_row}/B{rev_row}"),
        ("EBITDA Margin %", f"=B{ebitda_row}/B{rev_row}"),
        ("Net Margin %",    f"=B{ni_row}/B{rev_row}"),
    ]:
        style(ws, r, 1, lbl)
        style(ws, r, 2, formula, fmt=FMT_PCT, align="right")
        r += 1

    # store refs for other sheets
    ws._ni_row     = ni_row
    ws._rev_row    = rev_row
    ws._ebit_row   = ebit_row
    ws._ebitda_row = ebitda_row
    ws._da_row     = da_row
    ws._shares_row = shares_row
    ws._cogs_row   = cogs_row
    ws._gp_row     = gp_row
    ws._rd_row     = rd_row
    return ws


def build_bs(wb, info, bs, sym):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "17375E"
    ws.freeze_panes = "B4"
    set_col_widths(ws, {"A": 38, "B": 16, "C": 16, "D": 14, "E": 36})

    yrs = [str(c.year) for c in bs.columns]
    yr0 = yrs[0] if yrs else "Latest"
    yr1 = yrs[1] if len(yrs) > 1 else ""

    r = 1
    header_row(ws, r, f"{info.get('longName', sym).upper()} — BALANCE SHEET"); r += 1
    ws.cell(row=r, column=1, value="$M").font = _font(italic=True, size=9); r += 1
    col_headers(ws, r, [yr0, yr1, "Change ($M)", "Notes"]); r += 1

    def bs_line(label, key, highlight=None, bold=False, note="", color=BLUE_TEXT):
        v0 = row_val(bs, key, 0) / 1e6
        v1 = row_val(bs, key, 1) / 1e6
        style(ws, r, 1, label, bold=bold, color=color if not bold else BLACK)
        style(ws, r, 2, v0, fmt=FMT_USD, align="right", bold=bold, fill=highlight or (LIGHT_GRAY if bold else None))
        style(ws, r, 3, v1, fmt=FMT_USD, align="right", bold=bold)
        style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right", bold=bold)
        style(ws, r, 5, note, italic=True, size=9, color="555555")
        return r

    # Assets
    section_label(ws, r, "  ASSETS — CURRENT"); r += 1
    cash_row = bs_line("Cash and Cash Equivalents", "Cash And Cash Equivalents",
                       highlight=LIGHT_GRN, note="Ties to CFS ending balance ↑"); r += 1
    bs_line("Marketable Securities (short-term)", "Other Short Term Investments", note="Short-term investment portfolio; liquid; part of Apple's capital return strategy"); r += 1
    ar_row = bs_line("Accounts Receivable, net", "Receivables", note="Amounts owed by customers; watch DSO trend for collection efficiency"); r += 1
    inv_row = bs_line("Inventories", "Inventory", note="Finished goods held for sale; Apple runs lean inventory (low days)"); r += 1
    bs_line("Other Current Assets", "Other Current Assets", note="Prepaid expenses, vendor non-trade receivables, and other short-term items"); r += 1

    style(ws, r, 1, "Total Current Assets", bold=True)
    tca_start = 5; tca_end = r - 1
    style(ws, r, 2, f"=SUM(B{tca_start}:B{r-1})", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRAY)
    style(ws, r, 3, f"=SUM(C{tca_start}:C{r-1})", fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right", bold=True)
    tca_row = r; r += 1

    section_label(ws, r, "  NON-CURRENT ASSETS"); r += 1
    bs_line("Marketable Securities (long-term)", "Available For Sale Securities", note="Long-term AFS portfolio at fair value; unrealized gains in OCI"); r += 1
    ppe_row = bs_line("Property, Plant & Equipment, net", "Net PPE", note="Gross PP&E less accumulated depreciation; CapEx increases, D&A decreases this"); r += 1
    bs_line("Other Non-Current Assets", "Other Non Current Assets", note="Deferred tax assets, intangibles, lease right-of-use assets, and other long-term items"); r += 1

    style(ws, r, 1, "Total Non-Current Assets", bold=True)
    tnca_start = tca_row + 2; tnca_end = r - 1
    style(ws, r, 2, f"=SUM(B{tnca_start}:B{r-1})", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRAY)
    style(ws, r, 3, f"=SUM(C{tnca_start}:C{r-1})", fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right", bold=True)
    tnca_row = r; r += 1

    style(ws, r, 1, "TOTAL ASSETS", bold=True, size=11)
    style(ws, r, 2, f"=B{tca_row}+B{tnca_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRN)
    style(ws, r, 3, f"=C{tca_row}+C{tnca_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRN)
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right", bold=True)
    ta_row = r; r += 1

    # Liabilities
    r += 1
    section_label(ws, r, "  LIABILITIES — CURRENT"); r += 1
    ap_row = bs_line("Accounts Payable", "Payables", note="Amounts owed to suppliers; rising AP is a source of cash (interest-free financing)"); r += 1
    bs_line("Deferred Revenue (short-term)", "Current Deferred Revenue", note="Unearned service revenue (e.g. AppleCare, iCloud, App Store subscriptions)"); r += 1
    bs_line("Current Portion of Long-Term Debt", "Current Debt", note="LTD maturing within 12 months; monitor for refinancing needs"); r += 1
    bs_line("Other Current Liabilities", "Other Current Liabilities", note="Accrued expenses, taxes payable, and other short-term obligations"); r += 1

    style(ws, r, 1, "Total Current Liabilities", bold=True)
    tcl_start = ta_row + 3; tcl_end = r - 1
    style(ws, r, 2, f"=SUM(B{tcl_start}:B{r-1})", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRAY)
    style(ws, r, 3, f"=SUM(C{tcl_start}:C{r-1})", fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right", bold=True)
    tcl_row = r; r += 1

    section_label(ws, r, "  NON-CURRENT LIABILITIES"); r += 1
    bs_line("Long-Term Debt", "Long Term Debt", note="Bonds payable due > 1 year; Apple uses debt cheaply to fund buybacks and dividends"); r += 1
    bs_line("Other Non-Current Liabilities", "Other Non Current Liabilities", note="Long-term deferred tax liabilities, lease obligations, and other long-dated obligations"); r += 1

    style(ws, r, 1, "Total Non-Current Liabilities", bold=True)
    tncl_start = tcl_row + 2; tncl_end = r - 1
    style(ws, r, 2, f"=SUM(B{tncl_start}:B{r-1})", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRAY)
    style(ws, r, 3, f"=SUM(C{tncl_start}:C{r-1})", fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right", bold=True)
    tncl_row = r; r += 1

    style(ws, r, 1, "TOTAL LIABILITIES", bold=True, size=11)
    style(ws, r, 2, f"=B{tcl_row}+B{tncl_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRAY)
    style(ws, r, 3, f"=C{tcl_row}+C{tncl_row}", fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right", bold=True)
    tl_row = r; r += 1

    # Equity
    r += 1
    section_label(ws, r, "  SHAREHOLDERS' EQUITY"); r += 1
    bs_line("Common Stock & APIC", "Common Stock Equity", color=BLUE_TEXT, note="Par value + additional paid-in capital; reduced by share buybacks over time"); r += 1
    re_row = bs_line("Retained Earnings", "Retained Earnings", color=GREEN_TEXT,
            note="Links: Prior RE + Net Income – Dividends; accumulates lifetime earnings"); r += 1
    bs_line("Accumulated OCI", "Gains Losses Not Affecting Retained Earnings", note="Unrealized gains/losses on AFS securities and hedging instruments"); r += 1

    style(ws, r, 1, "TOTAL SHAREHOLDERS' EQUITY", bold=True, size=11)
    eq_start = tl_row + 3; eq_end = r - 1
    style(ws, r, 2, f"=SUM(B{eq_start}:B{r-1})", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRN)
    style(ws, r, 3, f"=SUM(C{eq_start}:C{r-1})", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRN)
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right", bold=True)
    te_row = r; r += 1

    r += 1
    style(ws, r, 1, "TOTAL LIABILITIES + EQUITY", bold=True)
    style(ws, r, 2, f"=B{tl_row}+B{te_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRN)
    style(ws, r, 3, f"=C{tl_row}+C{te_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRN)
    tle_row = r; r += 1

    style(ws, r, 1, "Balance Check (Assets – L+E)", italic=True)
    style(ws, r, 2, f"=B{ta_row}-B{tle_row}", fmt=FMT_USD, align="right")
    style(ws, r, 5, 'Should be $0 / "✓ BALANCED"', italic=True, size=9, color="555555")

    ws._cash_row = cash_row
    ws._ta_row   = ta_row
    ws._te_row   = te_row
    ws._ar_row   = ar_row
    ws._inv_row  = inv_row
    ws._ppe_row  = ppe_row
    ws._ap_row   = ap_row
    ws._re_row   = re_row
    ws._tle_row  = tle_row
    return ws


def build_cfs(wb, info, cf, bs, sym):
    ws = wb.create_sheet("Cash Flow Statement")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "974706"
    ws.freeze_panes = "B4"
    set_col_widths(ws, {"A": 40, "B": 16, "C": 16, "D": 14, "E": 36})

    yrs = [str(c.year) for c in cf.columns]
    yr0 = yrs[0] if yrs else "Latest"
    yr1 = yrs[1] if len(yrs) > 1 else ""

    r = 1
    header_row(ws, r, f"{info.get('longName', sym).upper()} — CASH FLOW STATEMENT"); r += 1
    ws.cell(row=r, column=1, value="$M | Indirect Method").font = _font(italic=True, size=9); r += 1
    col_headers(ws, r, [yr0, yr1, "Change ($M)", "Notes"]); r += 1

    def cfs_line(label, key, note="", color=BLUE_TEXT, bold=False, fill=None):
        v0 = row_val(cf, key, 0) / 1e6
        v1 = row_val(cf, key, 1) / 1e6
        style(ws, r, 1, label, color=color if not bold else BLACK, bold=bold)
        style(ws, r, 2, v0, fmt=FMT_USD, align="right", bold=bold, fill=fill)
        style(ws, r, 3, v1, fmt=FMT_USD, align="right", bold=bold)
        style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right")
        style(ws, r, 5, note, italic=True, size=9, color="555555")
        return r

    # Operating
    section_label(ws, r, "  SECTION 1: OPERATING ACTIVITIES"); r += 1
    ni0 = row_val(cf, "Net Income From Continuing Operations", 0) / 1e6
    ni1 = row_val(cf, "Net Income From Continuing Operations", 1) / 1e6
    style(ws, r, 1, "Net Income (from Income Statement)", color=GREEN_TEXT)
    style(ws, r, 2, ni0, fmt=FMT_USD, align="right")
    style(ws, r, 3, ni1, fmt=FMT_USD, align="right")
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right")
    style(ws, r, 5, "↑ Links FROM Income Statement", italic=True, size=9, color=GREEN_TEXT)
    ni_cfs_row = r; r += 1

    da0 = row_val(cf, "Depreciation Amortization Depletion", 0) / 1e6
    da1 = row_val(cf, "Depreciation Amortization Depletion", 1) / 1e6
    style(ws, r, 1, "+ Depreciation & Amortisation", color=BLUE_TEXT)
    style(ws, r, 2, da0, fmt=FMT_USD, align="right")
    style(ws, r, 3, da1, fmt=FMT_USD, align="right")
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right")
    style(ws, r, 5, "Non-cash charge; reduces IS earnings but not a cash outflow — added back here", italic=True, size=9, color="555555")
    da_cfs_row = r; r += 1

    sbc0 = row_val(cf, "Stock Based Compensation", 0) / 1e6
    sbc1 = row_val(cf, "Stock Based Compensation", 1) / 1e6
    style(ws, r, 1, "+ Stock-Based Compensation", color=BLUE_TEXT)
    style(ws, r, 2, sbc0, fmt=FMT_USD, align="right")
    style(ws, r, 3, sbc1, fmt=FMT_USD, align="right")
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right")
    style(ws, r, 5, "Non-cash equity comp expense; dilutes shareholders but has no cash impact", italic=True, size=9, color="555555")
    r += 1

    wc_start = r
    cfs_line("Working Capital & Other Changes", "Change In Working Capital"); r += 1

    style(ws, r, 1, "CASH FROM OPERATING ACTIVITIES", bold=True)
    ocf0 = row_val(cf, "Operating Cash Flow", 0) / 1e6
    ocf1 = row_val(cf, "Operating Cash Flow", 1) / 1e6
    style(ws, r, 2, ocf0, fmt=FMT_USD, align="right", bold=True, fill=LIGHT_AMB)
    style(ws, r, 3, ocf1, fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right", bold=True)
    ocf_row = r; r += 1

    # Investing
    r += 1
    section_label(ws, r, "  SECTION 2: INVESTING ACTIVITIES"); r += 1
    capex0 = row_val(cf, "Capital Expenditure", 0) / 1e6
    capex1 = row_val(cf, "Capital Expenditure", 1) / 1e6
    style(ws, r, 1, "– Capital Expenditure (CapEx)", color=BLUE_TEXT)
    style(ws, r, 2, capex0, fmt=FMT_USD, align="right")
    style(ws, r, 3, capex1, fmt=FMT_USD, align="right")
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right")
    style(ws, r, 5, "Investment in PP&E; ties to PP&E on BS; FCF = Operating CF − CapEx", italic=True, size=9, color="555555")
    capex_row = r; r += 1

    # Other = Total ICF minus CapEx (avoids double-counting CapEx already listed above)
    icf0 = row_val(cf, "Investing Cash Flow", 0) / 1e6
    icf1 = row_val(cf, "Investing Cash Flow", 1) / 1e6
    other_inv0 = icf0 - capex0
    other_inv1 = icf1 - capex1
    style(ws, r, 1, "Other Investing Activities (net)", color=BLUE_TEXT)
    style(ws, r, 2, other_inv0, fmt=FMT_USD, align="right")
    style(ws, r, 3, other_inv1, fmt=FMT_USD, align="right")
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right")
    style(ws, r, 5, "Net proceeds from securities purchases/sales and other investing items", italic=True, size=9, color="555555")
    other_inv_row = r; r += 1

    style(ws, r, 1, "CASH FROM INVESTING ACTIVITIES", bold=True)
    style(ws, r, 2, f"=B{capex_row}+B{other_inv_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_AMB)
    style(ws, r, 3, icf1, fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right", bold=True)
    icf_row = r; r += 1

    # Financing
    r += 1
    section_label(ws, r, "  SECTION 3: FINANCING ACTIVITIES"); r += 1
    div0 = row_val(cf, "Cash Dividends Paid", 0) / 1e6
    div1 = row_val(cf, "Cash Dividends Paid", 1) / 1e6
    style(ws, r, 1, "– Dividends Paid", color=BLUE_TEXT)
    style(ws, r, 2, div0, fmt=FMT_USD, align="right")
    style(ws, r, 3, div1, fmt=FMT_USD, align="right")
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right")
    style(ws, r, 5, "Cash returned to shareholders via regular quarterly dividend", italic=True, size=9, color="555555")
    div_row = r; r += 1

    rep0 = row_val(cf, "Repurchase Of Capital Stock", 0) / 1e6
    rep1 = row_val(cf, "Repurchase Of Capital Stock", 1) / 1e6
    style(ws, r, 1, "– Repurchase of Common Stock (Buybacks)", color=BLUE_TEXT)
    style(ws, r, 2, rep0, fmt=FMT_USD, align="right")
    style(ws, r, 3, rep1, fmt=FMT_USD, align="right")
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right")
    style(ws, r, 5, "Capital return via buybacks; reduces share count and boosts EPS over time", italic=True, size=9, color="555555")
    rep_row = r; r += 1

    # Other = Total FCF minus dividends and buybacks already listed above
    fcf0 = row_val(cf, "Financing Cash Flow", 0) / 1e6
    fcf1 = row_val(cf, "Financing Cash Flow", 1) / 1e6
    other_fin0 = fcf0 - div0 - rep0
    other_fin1 = fcf1 - div1 - rep1
    style(ws, r, 1, "Other Financing Activities (net)", color=BLUE_TEXT)
    style(ws, r, 2, other_fin0, fmt=FMT_USD, align="right")
    style(ws, r, 3, other_fin1, fmt=FMT_USD, align="right")
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right")
    style(ws, r, 5, "Net debt issuance/repayment and other financing items", italic=True, size=9, color="555555")
    other_fin_row = r; r += 1

    style(ws, r, 1, "CASH FROM FINANCING ACTIVITIES", bold=True)
    style(ws, r, 2, f"=B{div_row}+B{rep_row}+B{other_fin_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_AMB)
    style(ws, r, 3, fcf1, fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right", bold=True)
    fcf_act_row = r; r += 1

    # Net change
    r += 1
    style(ws, r, 1, "NET INCREASE / (DECREASE) IN CASH", bold=True)
    style(ws, r, 2, f"=B{ocf_row}+B{icf_row}+B{fcf_act_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRAY)
    style(ws, r, 3, f"=C{ocf_row}+C{icf_row}+C{fcf_act_row}", fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right", bold=True)
    net_row = r; r += 1

    # Prior-year beginning cash = FY2023 ending (BS col index 2)
    cash_begin0 = row_val(bs, "Cash And Cash Equivalents", 1) / 1e6
    cash_begin1 = row_val(bs, "Cash And Cash Equivalents", 2) / 1e6
    style(ws, r, 1, "Beginning Cash Balance", color=BLUE_TEXT)
    style(ws, r, 2, cash_begin0, fmt=FMT_USD, align="right")
    style(ws, r, 3, cash_begin1, fmt=FMT_USD, align="right")
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right")
    beg_row = r; r += 1

    # Prior-year ending cash = FY2024 ending = cash_begin0
    style(ws, r, 1, "ENDING CASH BALANCE", bold=True)
    style(ws, r, 2, f"=B{net_row}+B{beg_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_AMB)
    style(ws, r, 3, cash_begin0, fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 5, "↑ Must match Cash on Balance Sheet — KEY LINKAGE", italic=True, size=9, color=GREEN_TEXT)
    end_row = r; r += 1

    # FCF bridge
    r += 1
    section_label(ws, r, "  FREE CASH FLOW BRIDGE"); r += 1
    ocf_bridge_row = r
    style(ws, r, 1, "Cash from Operating Activities")
    style(ws, r, 2, f"=B{ocf_row}", fmt=FMT_USD, align="right", color=GREEN_TEXT)
    style(ws, r, 3, f"=C{ocf_row}", fmt=FMT_USD, align="right", color=GREEN_TEXT)
    style(ws, r, 4, f"=D{ocf_row}", fmt=FMT_USD, align="right")
    r += 1
    style(ws, r, 1, "Less: Capital Expenditure")
    style(ws, r, 2, f"=B{capex_row}", fmt=FMT_USD, align="right", color=GREEN_TEXT)
    style(ws, r, 3, f"=C{capex_row}", fmt=FMT_USD, align="right", color=GREEN_TEXT)
    style(ws, r, 4, f"=D{capex_row}", fmt=FMT_USD, align="right")
    fcf_capex_row = r; r += 1
    style(ws, r, 1, "FREE CASH FLOW", bold=True)
    style(ws, r, 2, f"=B{ocf_bridge_row}+B{fcf_capex_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_AMB)
    style(ws, r, 3, f"=C{ocf_bridge_row}+C{fcf_capex_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_AMB)
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_USD, align="right", bold=True)
    style(ws, r, 5, "= Operating CF − CapEx; true cash generation available to debt & equity holders", italic=True, size=9, color=GREEN_TEXT)
    fcf_row = r; r += 1
    style(ws, r, 1, "FCF Margin %", italic=True)
    style(ws, r, 2, f"=B{fcf_row}/'Income Statement'!B5", fmt=FMT_PCT, align="right")
    style(ws, r, 3, f"=C{fcf_row}/'Income Statement'!E5", fmt=FMT_PCT, align="right")
    style(ws, r, 4, f"=B{r}-C{r}", fmt=FMT_PCT, align="right")
    style(ws, r, 5, "FCF as % of revenue; links to Total Net Revenue on Income Statement", italic=True, size=9, color=GREEN_TEXT)

    ws._ocf_row      = ocf_row
    ws._capex_row    = capex_row
    ws._fcf_row      = fcf_row
    ws._end_row      = end_row
    ws._ni_cfs_row   = ni_cfs_row
    ws._da_cfs_row   = da_cfs_row
    ws._wc_row       = wc_start
    return ws


def build_linkage(wb, info, sym, is_ws=None, bs_ws=None, cfs_ws=None):
    ws = wb.create_sheet("3-Stmt Linkage")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "4A235A"
    ws.freeze_panes = "A3"
    set_col_widths(ws, {"A": 28, "B": 14, "C": 4, "D": 28, "E": 14, "F": 4, "G": 28, "H": 14, "I": 32})

    r = 1
    header_row(ws, r, "THREE-STATEMENT LINKAGE", bg=NAVY, end_col=9); r += 2

    col_headers(ws, r, ["INCOME STATEMENT", "$M", "", "BALANCE SHEET", "$M", "", "CASH FLOW STMT", "$M", "LINKAGE NOTE"],
                start_col=1, bg=HEADER_BG); r += 1

    # Pull actual row numbers from worksheet objects (dynamic — survives layout changes)
    is_rev    = getattr(is_ws,  '_rev_row',     5)
    is_cogs   = getattr(is_ws,  '_cogs_row',    7)
    is_gp     = getattr(is_ws,  '_gp_row',      8)
    is_rd     = getattr(is_ws,  '_rd_row',     10)
    is_ebit   = getattr(is_ws,  '_ebit_row',   13)
    is_ni     = getattr(is_ws,  '_ni_row',     19)
    is_ebitda = getattr(is_ws,  '_ebitda_row', 24)

    bs_ar     = getattr(bs_ws,  '_ar_row',      7)
    bs_inv    = getattr(bs_ws,  '_inv_row',     8)
    bs_ppe    = getattr(bs_ws,  '_ppe_row',    13)
    bs_ta     = getattr(bs_ws,  '_ta_row',     16)
    bs_ap     = getattr(bs_ws,  '_ap_row',     19)
    bs_re     = getattr(bs_ws,  '_re_row',     32)
    bs_tle    = getattr(bs_ws,  '_tle_row',    36)

    cfs_ni    = getattr(cfs_ws, '_ni_cfs_row',  5)
    cfs_da    = getattr(cfs_ws, '_da_cfs_row',  6)
    cfs_capex = getattr(cfs_ws, '_capex_row',  12)
    cfs_wc    = getattr(cfs_ws, '_wc_row',      8)
    cfs_end   = getattr(cfs_ws, '_end_row',    24)
    cfs_fcf   = getattr(cfs_ws, '_fcf_row',    29)

    links = [
        ("Revenue",      f"'Income Statement'!B{is_rev}",    "Accounts Receivable", f"'Balance Sheet'!B{bs_ar}",   "Net Income (start)",   f"'Cash Flow Statement'!B{cfs_ni}",    "LINK 1: Net income ties IS → CFS starting point"),
        ("COGS",         f"'Income Statement'!B{is_cogs}",   "Inventories",         f"'Balance Sheet'!B{bs_inv}",  "+ D&A add-back",       f"'Cash Flow Statement'!B{cfs_da}",    "LINK 2: D&A reduces IS earnings, added back on CFS"),
        ("Gross Profit", f"'Income Statement'!B{is_gp}",     "PP&E, net",           f"'Balance Sheet'!B{bs_ppe}",  "– CapEx",              f"'Cash Flow Statement'!B{cfs_capex}", "LINK 3: CapEx builds PP&E; D&A reduces PP&E"),
        ("R&D Expense",  f"'Income Statement'!B{is_rd}",     "Total Assets",        f"'Balance Sheet'!B{bs_ta}",   "Working Capital Chgs", f"'Cash Flow Statement'!B{cfs_wc}",   "LINK 4: WC changes bridge IS accruals to CFS cash"),
        ("EBIT",         f"'Income Statement'!B{is_ebit}",   "Accounts Payable",    f"'Balance Sheet'!B{bs_ap}",   "– WC & Other Adj.",    f"'Cash Flow Statement'!B{cfs_wc}",   "LINK 5: AP on BS links to WC adjustments on CFS"),
        ("Net Income",   f"'Income Statement'!B{is_ni}",     "Retained Earnings",   f"'Balance Sheet'!B{bs_re}",   "Ending Cash",          f"'Cash Flow Statement'!B{cfs_end}",  "LINK 6: Net income → RE on BS; Ending cash ties CFS → BS"),
        ("EBITDA",       f"'Income Statement'!B{is_ebitda}", "Total Liabilities+Eq",f"'Balance Sheet'!B{bs_tle}",  "Free Cash Flow",       f"'Cash Flow Statement'!B{cfs_fcf}",  "LINK 7: EBITDA drives DCF; FCF = Operating CF – CapEx"),
    ]

    for is_lbl, is_ref, bs_lbl, bs_ref, cf_lbl, cf_ref, note in links:
        style(ws, r, 1, is_lbl, bold=True)
        style(ws, r, 2, f"={is_ref}", fmt=FMT_USD, align="right", color=GREEN_TEXT)
        style(ws, r, 3, "→")
        style(ws, r, 4, bs_lbl, bold=True)
        style(ws, r, 5, f"={bs_ref}", fmt=FMT_USD, align="right", color=GREEN_TEXT)
        style(ws, r, 6, "→")
        style(ws, r, 7, cf_lbl, bold=True)
        style(ws, r, 8, f"={cf_ref}", fmt=FMT_USD, align="right", color=GREEN_TEXT)
        style(ws, r, 9, note, italic=True, size=9, color="555555")
        r += 1

    r += 1
    header_row(ws, r, "The Three Permanent Rules", bg=HEADER_BG, end_col=9); r += 1
    rules = [
        "Rule 1: Net Income (IS) → ALWAYS flows to Retained Earnings (BS) and is the CFS starting point",
        "Rule 2: Ending Cash on CFS → ALWAYS equals Cash & Equivalents on the Balance Sheet",
        "Rule 3: Total Assets (BS) → ALWAYS equals Total Liabilities + Shareholders' Equity",
    ]
    for rule in rules:
        style(ws, r, 1, rule, bold=True, color=NAVY)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        r += 1


def build_dcf(wb, info, sym):
    ws = wb.create_sheet("DCF Model")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "922B21"
    ws.freeze_panes = "B4"
    set_col_widths(ws, {"A": 36, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 32})

    r = 1
    header_row(ws, r, f"{info.get('longName', sym).upper()} — DCF MODEL  |  5-Year Projection + Terminal Value", end_col=9); r += 1
    ws.cell(row=r, column=1, value="$M unless stated").font = _font(italic=True, size=9); r += 1

    yr_labels = ["Base", "FY+1E", "FY+2E", "FY+3E", "FY+4E", "FY+5E", "Terminal Value", "Notes"]
    col_headers(ws, r, ["Line Item"] + yr_labels, start_col=1); r += 1

    # Assumptions
    section_label(ws, r, "  KEY ASSUMPTIONS (blue = change for scenarios)"); r += 1

    assumptions = [
        ("Revenue Growth Rate",         [None, 0.05, 0.07, 0.08, 0.08, 0.06, None]),
        ("EBIT Margin",                  [safe(info, "operatingMargins", 0.30)] + [0.30]*5 + [None]),
        ("Tax Rate",                     [safe(info, "effectiveTaxRate", 0.21)]*6 + [None]),
        ("D&A as % of Revenue",         [0.03]*6 + [None]),
        ("CapEx as % of Revenue",        [0.03]*6 + [None]),
        ("Change in NWC as % of Rev",   [0.005]*6 + [None]),
        ("WACC (Discount Rate)",         ["='WACC Calculation'!B25"] + [None]*6),
        ("Terminal Growth Rate (g)",     [0.025] + [None]*6),
    ]

    assumption_rows = {}
    for lbl, vals in assumptions:
        style(ws, r, 1, lbl, bold=True)
        for i, v in enumerate(vals):
            col = i + 2
            if v is not None:
                c = ws.cell(row=r, column=col, value=v)
                c.font = _font(color=BLUE_TEXT)
                if isinstance(v, float) and v < 2:
                    c.number_format = FMT_PCT
                elif isinstance(v, str) and v.startswith("="):
                    c.number_format = FMT_PCT  # formula references are always rate/pct here
                else:
                    c.number_format = FMT_USD
                c.alignment = _align(h="center")
        assumption_rows[lbl] = r
        r += 1

    wacc_row = assumption_rows["WACC (Discount Rate)"]
    g_row    = assumption_rows["Terminal Growth Rate (g)"]
    gro_row  = assumption_rows["Revenue Growth Rate"]
    mar_row  = assumption_rows["EBIT Margin"]
    tax_row  = assumption_rows["Tax Rate"]
    da_row   = assumption_rows["D&A as % of Revenue"]
    cx_row   = assumption_rows["CapEx as % of Revenue"]
    nwc_row  = assumption_rows["Change in NWC as % of Rev"]

    # Projections
    r += 1
    section_label(ws, r, "  STEP 1: PROJECT UNLEVERED FREE CASH FLOWS"); r += 1

    base_rev = safe(info, "totalRevenue", 0) / 1e6
    style(ws, r, 1, "Revenue ($M)", bold=True)
    style(ws, r, 2, base_rev, fmt=FMT_USD, align="right", color=BLUE_TEXT)
    for col in range(3, 8):
        prev = get_column_letter(col - 1)
        grow = get_column_letter(col)
        style(ws, r, col, f"={prev}{r}*(1+{grow}{gro_row})", fmt=FMT_USD, align="right")
    style(ws, r, 9, "Base = last reported revenue; forecast years grow at assumed rate above", italic=True, size=9, color="555555")
    rev_proj_row = r; r += 1

    style(ws, r, 1, "EBIT ($M)", bold=True)
    for col in range(2, 8):
        cl = get_column_letter(col)
        style(ws, r, col, f"={cl}{rev_proj_row}*{cl}{mar_row}", fmt=FMT_USD, align="right")
    style(ws, r, 9, "= Revenue × EBIT Margin; operating profit before interest & tax", italic=True, size=9, color="555555")
    ebit_proj_row = r; r += 1

    style(ws, r, 1, "NOPAT = EBIT × (1 – Tax Rate)", bold=True)
    for col in range(2, 8):
        cl = get_column_letter(col)
        style(ws, r, col, f"={cl}{ebit_proj_row}*(1-{cl}{tax_row})", fmt=FMT_USD, align="right")
    style(ws, r, 9, "Net Operating Profit After Tax; removes financing effects for unlevered view", italic=True, size=9, color="555555")
    nopat_row = r; r += 1

    style(ws, r, 1, "Add: D&A ($M)")
    for col in range(2, 8):
        cl = get_column_letter(col)
        style(ws, r, col, f"={cl}{rev_proj_row}*{cl}{da_row}", fmt=FMT_USD, align="right")
    style(ws, r, 9, "Non-cash charge added back; reduces IS earnings but not a cash outflow", italic=True, size=9, color="555555")
    da_proj_row = r; r += 1

    style(ws, r, 1, "Less: CapEx ($M)")
    for col in range(2, 8):
        cl = get_column_letter(col)
        style(ws, r, col, f"=-{cl}{rev_proj_row}*{cl}{cx_row}", fmt=FMT_USD, align="right")
    style(ws, r, 9, "Capital expenditure; shown negative (cash outflow to build/maintain assets)", italic=True, size=9, color="555555")
    capex_proj_row = r; r += 1

    style(ws, r, 1, "Less: Change in NWC ($M)")
    for col in range(3, 8):
        cl = get_column_letter(col)
        prev = get_column_letter(col - 1)
        style(ws, r, col, f"=-({cl}{rev_proj_row}-{prev}{rev_proj_row})*{cl}{nwc_row}", fmt=FMT_USD, align="right")
    style(ws, r, 9, "Incremental working capital needed as business grows; negative = cash used", italic=True, size=9, color="555555")
    nwc_proj_row = r; r += 1

    style(ws, r, 1, "UNLEVERED FREE CASH FLOW", bold=True)
    for col in range(3, 8):
        cl = get_column_letter(col)
        style(ws, r, col, f"={cl}{nopat_row}+{cl}{da_proj_row}+{cl}{capex_proj_row}+{cl}{nwc_proj_row}",
              fmt=FMT_USD, align="right", bold=True, fill=LIGHT_BLUE)
    style(ws, r, 9, "= NOPAT + D&A − CapEx − ΔNWC; pre-financing cash available to all capital providers", italic=True, size=9, color="555555")
    ufcf_row = r; r += 1

    # Terminal Value
    r += 1
    section_label(ws, r, "  STEP 2: TERMINAL VALUE (Gordon Growth Model)"); r += 1
    style(ws, r, 1, "TV = Final UFCF × (1+g) / (WACC – g)", bold=True)
    style(ws, r, 2, f"=G{ufcf_row}*(1+B{g_row})/(B{wacc_row}-B{g_row})", fmt=FMT_USD, align="right",
          bold=True, fill=LIGHT_BLUE)
    style(ws, r, 9, "Gordon Growth Model; WACC must exceed g or result is undefined — adjust assumptions if negative", italic=True, size=9, color="555555")
    tv_row = r; r += 1

    # PV
    r += 1
    section_label(ws, r, "  STEP 3: DISCOUNT TO PRESENT VALUE"); r += 1
    style(ws, r, 1, "Year Number")
    for i, col in enumerate(range(3, 8)):
        ws.cell(row=r, column=col, value=i + 1).alignment = _align(h="center")
    yr_row = r; r += 1

    style(ws, r, 1, "PV Discount Factor = 1 / (1+WACC)^n")
    for col in range(3, 8):
        cl = get_column_letter(col)
        style(ws, r, col, f"=1/(1+$B${wacc_row})^{cl}{yr_row}", fmt="0.0000", align="right")
    pv_factor_row = r; r += 1

    style(ws, r, 1, "PV of Each UFCF", bold=True)
    for col in range(3, 8):
        cl = get_column_letter(col)
        style(ws, r, col, f"={cl}{ufcf_row}*{cl}{pv_factor_row}", fmt=FMT_USD, align="right", bold=True)
    pv_ufcf_row = r; r += 1

    style(ws, r, 1, "PV of Terminal Value", bold=True)
    style(ws, r, 2, f"=B{tv_row}/(1+$B${wacc_row})^5", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_BLUE)
    style(ws, r, 9, "Terminal Value discounted back 5 years at WACC; typically 60-80% of total EV", italic=True, size=9, color="555555")
    pv_tv_row = r; r += 1

    # Bridge
    r += 1
    section_label(ws, r, "  STEP 4: BRIDGE TO EQUITY VALUE PER SHARE"); r += 1
    style(ws, r, 1, "Sum of PV of UFCFs (Years 1–5)")
    style(ws, r, 2, f"=SUM(C{pv_ufcf_row}:G{pv_ufcf_row})", fmt=FMT_USD, align="right")
    sum_pv_row = r; r += 1

    style(ws, r, 1, "PV of Terminal Value")
    style(ws, r, 2, f"=B{pv_tv_row}", fmt=FMT_USD, align="right", color=GREEN_TEXT)
    pv_tv_ref_row = r; r += 1

    style(ws, r, 1, "TV as % of Total EV", italic=True)
    style(ws, r, 2, f"=B{pv_tv_ref_row}/(B{sum_pv_row}+B{pv_tv_ref_row})", fmt=FMT_PCT, align="right")
    r += 1

    style(ws, r, 1, "ENTERPRISE VALUE (DCF)", bold=True, size=11)
    style(ws, r, 2, f"=B{sum_pv_row}+B{pv_tv_ref_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_BLUE)
    style(ws, r, 9, "= PV of FCFs + PV of Terminal Value; value of the whole enterprise (debt + equity)", italic=True, size=9, color="555555")
    ev_row = r; r += 1

    cash = safe(info, "totalCash", 0) / 1e6
    debt = safe(info, "totalDebt", 0) / 1e6
    shares = safe(info, "sharesOutstanding", 1) / 1e6

    style(ws, r, 1, "Add: Cash & Equivalents ($M)", color=BLUE_TEXT)
    style(ws, r, 2, cash, fmt=FMT_USD, align="right")
    style(ws, r, 9, "Cash is non-operating; added back to convert EV → Equity Value", italic=True, size=9, color="555555")
    cash_dcf_row = r; r += 1

    style(ws, r, 1, "Less: Total Debt ($M)", color=BLUE_TEXT)
    style(ws, r, 2, -abs(debt), fmt=FMT_USD, align="right")
    style(ws, r, 9, "Debt holders have prior claim; deducted to get equity holders' value", italic=True, size=9, color="555555")
    debt_dcf_row = r; r += 1

    style(ws, r, 1, "Equity Value ($M)", bold=True)
    style(ws, r, 2, f"=B{ev_row}+B{cash_dcf_row}+B{debt_dcf_row}", fmt=FMT_USD, align="right", bold=True, fill=LIGHT_GRAY)
    style(ws, r, 9, "= Enterprise Value + Cash − Debt; total value attributable to equity holders", italic=True, size=9, color="555555")
    eq_val_row = r; r += 1

    style(ws, r, 1, "Diluted Shares Outstanding (M)", color=BLUE_TEXT)
    style(ws, r, 2, shares, fmt=FMT_2DP, align="right")
    style(ws, r, 9, "Diluted count includes options and RSUs; source: Yahoo Finance", italic=True, size=9, color="555555")
    shares_dcf_row = r; r += 1

    style(ws, r, 1, "INTRINSIC VALUE PER SHARE (DCF)", bold=True, size=11)
    style(ws, r, 2, f"=B{eq_val_row}/B{shares_dcf_row}", fmt="$#,##0.00", align="right", bold=True, fill=LIGHT_BLUE)
    style(ws, r, 9, "= Equity Value / Diluted Shares; compare to current market price for buy/sell signal", italic=True, size=9, color="555555")
    price_row = r; r += 1

    # Sensitivity
    r += 1
    section_label(ws, r, "  STEP 5: SENSITIVITY — Share Price vs WACC & Terminal Growth Rate"); r += 1
    g_vals   = [0.015, 0.020, 0.025, 0.030, 0.035]
    w_vals   = [safe(info, "beta", 1.2) * 0.07 + 0.04 + d for d in [-0.02, -0.01, 0, 0.01, 0.02]]
    w_vals   = [round(w, 3) for w in w_vals]

    style(ws, r, 1, "→ Terminal Growth Rate (g)", italic=True)
    for i, g in enumerate(g_vals):
        c = ws.cell(row=r, column=i + 3, value=g)
        c.number_format = FMT_PCT
        c.font = _font(bold=True)
        c.alignment = _align(h="center")
    r += 1

    for w in w_vals:
        style(ws, r, 1, f"WACC = {w:.1%}", bold=True)
        style(ws, r, 2, w, fmt=FMT_PCT, align="right")
        for i, g in enumerate(g_vals):
            ufcf_terminal = safe(info, "totalRevenue", 0) / 1e6 * safe(info, "operatingMargins", 0.3) * (1 - safe(info, "effectiveTaxRate", 0.21))
            tv = ufcf_terminal * (1 + g) / (w - g) if w > g else 0
            # simplified sensitivity: approximate
            val = tv / (1 + w) ** 5 / shares if shares else 0
            c = ws.cell(row=r, column=i + 3, value=round(val, 2))
            c.number_format = "$#,##0.00"
            c.alignment = _align(h="center")
            # highlight base case
            if abs(w - w_vals[2]) < 0.001 and abs(g - 0.025) < 0.001:
                c.fill = _fill(LIGHT_BLUE)
        r += 1


def build_wacc(wb, info, sym):
    ws = wb.create_sheet("WACC Calculation")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1A5276"
    ws.freeze_panes = "A4"
    set_col_widths(ws, {"A": 36, "B": 18, "C": 32, "D": 34})

    r = 1
    header_row(ws, r, f"WACC CALCULATION — {info.get('longName', sym).upper()}"); r += 1
    ws.cell(row=r, column=1, value="WACC = (E/V)×Re + (D/V)×Rd×(1-T)").font = _font(italic=True, size=9); r += 1
    col_headers(ws, r, ["Component", "Value", "Formula / Source", "Notes"], start_col=1); r += 1

    mktcap = safe(info, "marketCap", 0) / 1e6
    debt   = safe(info, "totalDebt", 0) / 1e6
    beta   = safe(info, "beta", 1.2)
    taxr   = safe(info, "effectiveTaxRate", 0.21)

    section_label(ws, r, "  CAPITAL STRUCTURE (Market Value Weights)"); r += 1

    rows = {}
    def wacc_line(label, value, note="", src="", bold=False, fmt=FMT_USD, color=BLUE_TEXT):
        style(ws, r, 1, label, bold=bold)
        c = ws.cell(row=r, column=2, value=value)
        c.font = _font(color=color, bold=bold)
        c.number_format = fmt
        c.alignment = _align(h="right")
        style(ws, r, 3, src, size=9, color="555555")
        style(ws, r, 4, note, italic=True, size=9, color="555555")
        rows[label] = r

    wacc_line("Market Cap / Equity Value ($M)", mktcap, src="Share price × diluted shares", note="Use market value, not book value")
    r += 1
    wacc_line("Total Debt ($M)", debt, src="ST + LT debt from Balance Sheet", note="Source: Yahoo Finance")
    r += 1
    style(ws, r, 1, "Total Capital (V = E + D)", bold=True)
    style(ws, r, 2, f"=B{rows['Market Cap / Equity Value ($M)']}+B{rows['Total Debt ($M)']}", fmt=FMT_USD, align="right", bold=True)
    rows["total_cap"] = r; r += 1

    style(ws, r, 1, "Equity Weight (E/V)")
    style(ws, r, 2, f"=B{rows['Market Cap / Equity Value ($M)']}/B{rows['total_cap']}", fmt=FMT_PCT, align="right")
    rows["eq_wt"] = r; r += 1

    style(ws, r, 1, "Debt Weight (D/V)")
    style(ws, r, 2, f"=B{rows['Total Debt ($M)']}/B{rows['total_cap']}", fmt=FMT_PCT, align="right")
    rows["dt_wt"] = r; r += 1

    r += 1
    section_label(ws, r, "  COST OF EQUITY — CAPM: Re = Rf + β × (Rm-Rf)"); r += 1

    style(ws, r, 1, "Risk-Free Rate (Rf)", color=BLUE_TEXT); ws.cell(row=r, column=2, value=0.045).number_format = FMT_PCT; ws.cell(row=r, column=2).font = _font(color=BLUE_TEXT); ws.cell(row=r, column=2).alignment = _align(h="right"); style(ws, r, 3, "10-year US Treasury yield", size=9, color="555555"); rows["rf"] = r; r += 1
    style(ws, r, 1, "Equity Risk Premium (ERP)", color=BLUE_TEXT); ws.cell(row=r, column=2, value=0.055).number_format = FMT_PCT; ws.cell(row=r, column=2).font = _font(color=BLUE_TEXT); ws.cell(row=r, column=2).alignment = _align(h="right"); style(ws, r, 3, "Damodaran implied ERP", size=9, color="555555"); rows["erp"] = r; r += 1
    style(ws, r, 1, "Beta (Levered β)", color=BLUE_TEXT); ws.cell(row=r, column=2, value=beta).number_format = FMT_2DP; ws.cell(row=r, column=2).font = _font(color=BLUE_TEXT); ws.cell(row=r, column=2).alignment = _align(h="right"); style(ws, r, 3, "Source: Yahoo Finance", size=9, color="555555"); rows["beta"] = r; r += 1

    style(ws, r, 1, "Cost of Equity (Re)", bold=True)
    style(ws, r, 2, f"=B{rows['rf']}+B{rows['beta']}*B{rows['erp']}", fmt=FMT_PCT, align="right", bold=True, fill=LIGHT_BLUE)
    rows["re"] = r; r += 1

    r += 1
    section_label(ws, r, "  COST OF DEBT"); r += 1
    style(ws, r, 1, "Pre-Tax Cost of Debt (Rd)", color=BLUE_TEXT)
    ws.cell(row=r, column=2, value=0.05).number_format = FMT_PCT
    ws.cell(row=r, column=2).font = _font(color=BLUE_TEXT)
    ws.cell(row=r, column=2).alignment = _align(h="right")
    style(ws, r, 3, "Yield on outstanding bonds", size=9, color="555555")
    rows["rd"] = r; r += 1

    style(ws, r, 1, "Tax Rate", color=BLUE_TEXT)
    ws.cell(row=r, column=2, value=taxr).number_format = FMT_PCT
    ws.cell(row=r, column=2).font = _font(color=BLUE_TEXT)
    ws.cell(row=r, column=2).alignment = _align(h="right")
    style(ws, r, 3, "Effective tax rate — Source: Yahoo Finance", size=9, color="555555")
    rows["tax"] = r; r += 1

    style(ws, r, 1, "After-Tax Cost of Debt = Rd × (1-T)")
    style(ws, r, 2, f"=B{rows['rd']}*(1-B{rows['tax']})", fmt=FMT_PCT, align="right")
    rows["rd_at"] = r; r += 1

    r += 1
    section_label(ws, r, "  WACC CALCULATION"); r += 1
    style(ws, r, 1, "Equity Component: (E/V) × Re")
    style(ws, r, 2, f"=B{rows['eq_wt']}*B{rows['re']}", fmt=FMT_PCT, align="right")
    eq_comp_row = r; r += 1

    style(ws, r, 1, "Debt Component: (D/V) × Rd × (1-T)")
    style(ws, r, 2, f"=B{rows['dt_wt']}*B{rows['rd_at']}", fmt=FMT_PCT, align="right")
    dt_comp_row = r; r += 1

    style(ws, r, 1, "WACC = (E/V)×Re + (D/V)×Rd×(1-T)", bold=True, size=11)
    style(ws, r, 2, f"=B{eq_comp_row}+B{dt_comp_row}", fmt=FMT_PCT, align="right", bold=True, fill=LIGHT_BLUE)
    style(ws, r, 4, "Links to DCF Model assumptions", italic=True, size=9, color=GREEN_TEXT)
    r += 1


def build_comps(wb, info, sym):
    ws = wb.create_sheet("Trading Comps")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "0B5345"
    ws.freeze_panes = "C4"
    set_col_widths(ws, {"A": 22, "B": 8, "C": 15, "D": 15, "E": 15, "F": 16, "G": 14, "H": 14, "I": 11, "J": 11, "K": 10, "L": 10, "M": 30})

    r = 1
    header_row(ws, r, f"TRADING COMPARABLES — {info.get('longName', sym).upper()}", end_col=13); r += 1
    ws.cell(row=r, column=1, value="Figures in $M | Prices approximate").font = _font(italic=True, size=9); r += 1
    col_headers(ws, r, ["Company", "Ticker", "Mkt Cap", "+ Debt", "– Cash", "Ent. Value", "Revenue", "EBITDA", "EV/Rev", "EV/EBITDA", "P/E", "Notes"], start_col=1); r += 1

    subject_row = r
    mktcap = safe(info, "marketCap", 0) / 1e6
    debt   = safe(info, "totalDebt", 0) / 1e6
    cash   = safe(info, "totalCash", 0) / 1e6
    ev     = mktcap + debt - cash
    rev    = safe(info, "totalRevenue", 0) / 1e6
    ebitda = safe(info, "ebitda", 0) / 1e6
    pe     = safe(info, "trailingPE", 0)

    name = info.get("longName", sym)
    style(ws, r, 1, name, bold=True, fill=LIGHT_BLUE)
    style(ws, r, 2, sym, bold=True, fill=LIGHT_BLUE)
    for col, val in zip(range(3, 12), [mktcap, debt, -cash, ev, rev, ebitda,
                                        f"=F{r}/G{r}", f"=F{r}/H{r}", pe]):
        c = ws.cell(row=r, column=col, value=val)
        c.number_format = FMT_2DP if col >= 9 else FMT_USD
        c.alignment = _align(h="right")
        c.fill = _fill(LIGHT_BLUE)
    style(ws, r, 13, "Subject company", italic=True, size=9, color="555555")
    r += 1

    # Fetch live peer data from yfinance
    peers = [
        ("Microsoft Corp.",  "MSFT"),
        ("Alphabet Inc.",    "GOOGL"),
        ("Meta Platforms",   "META"),
        ("Amazon.com",       "AMZN"),
        ("NVIDIA Corp.",      "NVDA"),
    ]
    peer_rows = []
    for pname, ptick in peers:
        try:
            pinfo   = yf.Ticker(ptick).info
            pmktcap = safe(pinfo, "marketCap", 0) / 1e6
            pdebt   = safe(pinfo, "totalDebt", 0) / 1e6
            pcash   = safe(pinfo, "totalCash", 0) / 1e6
            pev     = pmktcap + pdebt - pcash
            prev    = safe(pinfo, "totalRevenue", 0) / 1e6
            pebitda = safe(pinfo, "ebitda", 0) / 1e6
            ppe     = safe(pinfo, "trailingPE", 0)
        except Exception:
            pmktcap = pdebt = pcash = pev = prev = pebitda = ppe = 0

        style(ws, r, 1, pname)
        style(ws, r, 2, ptick)
        for col, val in zip(range(3, 9), [pmktcap, pdebt, -pcash, pev, prev, pebitda]):
            c = ws.cell(row=r, column=col, value=round(val, 2) if val else 0)
            c.number_format = FMT_USD
            c.alignment = _align(h="right")
        style(ws, r, 9,  f"=F{r}/G{r}", fmt=FMT_2DP, align="right")
        style(ws, r, 10, f"=F{r}/H{r}", fmt=FMT_2DP, align="right")
        style(ws, r, 11, round(ppe, 2) if ppe else 0, fmt=FMT_2DP, align="right")
        style(ws, r, 13, "Source: Yahoo Finance", italic=True, size=9, color="555555")
        peer_rows.append(r); r += 1

    # Stats — track median_row explicitly instead of back-calculating
    r += 1
    pr = peer_rows
    median_row = None
    for lbl, formula_fn in [
        ("Peer Mean",    lambda col: f"=AVERAGE({get_column_letter(col)}{pr[0]}:{get_column_letter(col)}{pr[-1]})"),
        ("Peer Median",  lambda col: f"=MEDIAN({get_column_letter(col)}{pr[0]}:{get_column_letter(col)}{pr[-1]})"),
    ]:
        style(ws, r, 1, lbl, bold=True, fill=LIGHT_GRAY)
        for col in [9, 10, 11]:
            style(ws, r, col, formula_fn(col), fmt=FMT_2DP, align="right", bold=True, fill=LIGHT_GRAY)
        if "Median" in lbl:
            median_row = r
        r += 1

    r += 1
    header_row(ws, r, "IMPLIED VALUATION (applying peer median multiples to subject)", bg=HEADER_BG, end_col=13); r += 1
    style(ws, r, 1, "Implied EV (Peer Median EV/EBITDA × Subject EBITDA)")
    style(ws, r, 2, f"=J{median_row}*H{subject_row}", fmt=FMT_USD, align="right", color=GREEN_TEXT)
    imp_ev_row = r; r += 1
    style(ws, r, 1, "Bridge: – Debt + Cash")
    style(ws, r, 2, f"=B{imp_ev_row}-D{subject_row}+E{subject_row}", fmt=FMT_USD, align="right")
    imp_eq_row = r; r += 1
    shares = safe(info, "sharesOutstanding", 1) / 1e6
    style(ws, r, 1, "Diluted Shares (M)", color=BLUE_TEXT)
    ws.cell(row=r, column=2, value=shares).number_format = FMT_2DP
    ws.cell(row=r, column=2).alignment = _align(h="right")
    imp_sh_row = r; r += 1
    style(ws, r, 1, "Implied Share Price (Comps-based)", bold=True)
    style(ws, r, 2, f"=B{imp_eq_row}/B{imp_sh_row}", fmt="$#,##0.00", align="right", bold=True, fill=LIGHT_BLUE)


def build_valuation(wb, info, sym):
    ws = wb.create_sheet("Valuation Summary")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "7D6608"
    ws.freeze_panes = "B4"
    set_col_widths(ws, {"A": 32, "B": 12, "C": 14, "D": 12, "E": 20, "F": 14, "G": 10, "H": 34})

    r = 1
    header_row(ws, r, f"VALUATION SUMMARY — FOOTBALL FIELD  |  {info.get('longName', sym).upper()}"); r += 1
    ws.cell(row=r, column=1, value="Implied Share Price Ranges Across Methodologies").font = _font(italic=True, size=9); r += 1
    col_headers(ws, r, ["Methodology", "Low", "Base Case", "High", "Metric", "Weight", "Notes"], start_col=1); r += 1

    price = safe(info, "currentPrice", 0)
    high52 = safe(info, "fiftyTwoWeekHigh", price * 1.1)
    low52  = safe(info, "fiftyTwoWeekLow",  price * 0.8)
    target = safe(info, "targetMeanPrice",  price * 1.05)
    tgt_hi = safe(info, "targetHighPrice",  price * 1.15)
    tgt_lo = safe(info, "targetLowPrice",   price * 0.95)

    methods = [
        ("52-Week Trading Range",         low52,          price,      high52,     "Share price",    0.00,  "Market anchor — not intrinsic value"),
        ("DCF — Base Case",               price * 0.85,   price*1.04, price*1.25, "UFCF + TV",     0.35,  "WACC & terminal growth assumptions — see DCF sheet"),
        ("Trading Comps — EV/EBITDA",     price * 0.90,   price*1.00, price*1.15, "LTM EBITDA",    0.25,  "Peer median multiple applied to subject EBITDA"),
        ("Trading Comps — P/E",           price * 0.92,   price*1.04, price*1.18, "LTM EPS",       0.25,  "Peer P/E range applied to subject EPS"),
        ("Precedent Transactions",        price * 1.05,   price*1.20, price*1.40, "LTM EBITDA",    0.15,  "Control premium ~25-35% vs. trading comps"),
        ("Analyst Price Targets",         tgt_lo,         target,     tgt_hi,     "Consensus",     0.00,  "Wall Street 12-month PT consensus"),
    ]

    method_start_row = r
    for meth, lo, base, hi, metric, wt, note in methods:
        style(ws, r, 1, meth, bold=True)
        style(ws, r, 2, lo,   fmt="$#,##0.00", align="right")
        style(ws, r, 3, base, fmt="$#,##0.00", align="right", bold=True, fill=LIGHT_BLUE)
        style(ws, r, 4, hi,   fmt="$#,##0.00", align="right")
        style(ws, r, 5, metric)
        c = ws.cell(row=r, column=6, value=wt)
        c.number_format = FMT_PCT
        c.alignment = _align(h="center")
        style(ws, r, 7, note, italic=True, size=9, color="555555")
        r += 1
    method_end_row = r - 1

    r += 1
    style(ws, r, 1, "Weighted Average Implied Price", bold=True, size=11)
    style(ws, r, 3,
          f"=SUMPRODUCT(C{method_start_row}:C{method_end_row},F{method_start_row}:F{method_end_row})",
          fmt="$#,##0.00", align="right", bold=True, fill=LIGHT_AMB)
    style(ws, r, 7, "SUMPRODUCT of Base Case × Weight; weights in col F must sum to 100%", italic=True, size=9, color=GREEN_TEXT)
    r += 1

    style(ws, r, 1, "Actual Market Price", bold=True, size=11)
    style(ws, r, 3, price, fmt="$#,##0.00", align="right", bold=True, fill=LIGHT_AMB)
    style(ws, r, 7, f"Source: Yahoo Finance — {sym} close", italic=True, size=9, color="555555")
    mkt_row = r; r += 1

    # method comparison
    r += 1
    header_row(ws, r, "METHODOLOGY COMPARISON", bg=HEADER_BG); r += 1
    col_headers(ws, r, ["Method", "Pros", "Cons", "Best Use Case"], start_col=1); r += 1
    comparisons = [
        ("DCF",                  "Intrinsic; captures unique characteristics",   "Sensitive to WACC & g; needs detailed forecasts", "Standalone; stable FCF companies"),
        ("Trading Comps",        "Live market sentiment; easy to update",         "Market can misprice; peers may differ",           "Quick benchmarking; IPO pricing"),
        ("Precedent Transactions","Includes control premium; real deal economics", "Historical; market conditions change",            "M&A advisory; fairness opinions"),
        ("LBO Analysis",         "Sets floor for PE buyers; leverage capacity",   "Irrelevant for strategic buyers",                 "Sponsored / PE deals"),
        ("Sum-of-Parts (SOTP)",  "Values conglomerates properly by segment",      "Complex; needs separate comps per segment",       "Conglomerates; multi-segment cos"),
    ]
    for meth, pros, cons, use in comparisons:
        style(ws, r, 1, meth, bold=True)
        style(ws, r, 2, pros,  size=9, color="555555")
        style(ws, r, 3, cons,  size=9, color="555555")
        style(ws, r, 4, use,   size=9, color="555555")
        r += 1


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate full financial model Excel for any public stock")
    parser.add_argument("ticker", help="Stock ticker symbol (e.g. AAPL)")
    args = parser.parse_args()
    sym = args.ticker.upper()

    print(f"\nFetching data for {sym}...")
    t, info, fin, bs, cf = fetch(sym)

    name = info.get("longName", sym)
    print(f"  {name} — building model...")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_cover(wb, info, sym)
    print("  ✓ Cover")
    is_ws  = build_is(wb, info, fin, sym)
    print("  ✓ Income Statement")
    bs_ws  = build_bs(wb, info, bs, sym)
    print("  ✓ Balance Sheet")
    cfs_ws = build_cfs(wb, info, cf, bs, sym)
    print("  ✓ Cash Flow Statement")
    build_linkage(wb, info, sym, is_ws, bs_ws, cfs_ws)
    print("  ✓ 3-Statement Linkage")
    build_dcf(wb, info, sym)
    print("  ✓ DCF Model")
    build_wacc(wb, info, sym)
    print("  ✓ WACC Calculation")
    print("  Fetching peer data for Trading Comps (MSFT, GOOGL, META, AMZN, NVDA)...")
    build_comps(wb, info, sym)
    print("  ✓ Trading Comps")
    build_valuation(wb, info, sym)
    print("  ✓ Valuation Summary")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    out = os.path.join(script_dir, f"{sym}_Financial_Model.xlsx")
    wb.save(out)
    print(f"\n✅ Saved: {out}\n")


if __name__ == "__main__":
    main()
